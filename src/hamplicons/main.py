from collections import defaultdict, namedtuple
from json import dumps as json_dumps
from multiprocessing import cpu_count, Pool
from shutil import which
from datetime import datetime
from pathlib import Path
from argparse import ArgumentParser, ArgumentDefaultsHelpFormatter
from gzip import open as gzopen
from itertools import combinations
from subprocess import Popen
import logging
import os
import sys
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import coloredlogs

__version__ = "4.2.0"

# Input constants
# Expected format/extension of input files
READ_FORMAT = ".fastq.gz"
# Minimum read length, used to exclude primer dimers, etc.
MIN_READ_LENGTH = 50

# minimum number of reads a target must have to be reported
TARGET_TOTAL_THRESHOLD = 100
# minimum number of reads a specific target variation must have
TARGET_LOCAL_THRESHOLD_READS = 10
# minimum percentage of reads observed with a specific indel to be reported
TARGET_LOCAL_THRESHOLD_PERCENTAGE = 0.05

# Unmatched reporting constants
# report significant products with >= this reads
SIGNIFICANT_UNMATCHED_THRESHOLD_READS = 500
# require these products to be at least this long, to exclude probable primer-dimers
SIGNIFICANT_UNMATCHED_THRESHOLD_LENGTH = 50

# Barcode matching constants
# bp in start/end that must match to categorize fastq reads as a specific amplicon
BARSIZE = 30
# when performing additional matching
REDUCED_BAR = 15

# Other constants
# Size of regions compared between PCR products to avoid too similar sequences
WORRY_ZONE = 40


TargetSequence = namedtuple("TargetSequence", ["name", "seq", "len", "barcode_5p", "barcode_3p", "ham_5p", "ham_3p"],)


def run_command(command, **kwargs):
    proc = Popen(command, **kwargs)
    return proc.wait()


def setup_logging(level=logging.INFO):
    fmt = "%(asctime)s - %(message)s"
    if not sys.stdout.isatty():
        fmt = "%(asctime)s - %(levelname)s - %(message)s"

    coloredlogs.install(level=level, fmt=fmt, datefmt="%H:%M:%S", level_styles={
        "info": {"color": "white"}, "critical": {"color": "red", "bold": True}, "verbose": {"color": "blue"},
        "error": {"color": "red"}, "debug": {"color": "green"}, "warning": {"color": "yellow"},},
    )

    return logging.getLogger("main")


def setup_file_logging(filename, level=logging.INFO):
    # set up a global warning logger
    formatter = logging.Formatter("%(asctime)s : %(levelname)s : %(message)s")
    handler = logging.FileHandler(filename, mode="w")
    handler.setFormatter(formatter)
    handler.setLevel(level)
    logger = logging.getLogger("")
    logger.addHandler(handler)


def convert_index_to_well(index, rows=8, columns=12):  # 1 based index
    thisrow = (index - 1) // columns
    thisplate = thisrow // rows
    thiscolumn = index - thisrow * columns
    thisrow -= rows * thisplate
    return "{}{}_p{}".format(chr(65 + thisrow), thiscolumn, thisplate)


def hamming(seq1, seq2):
    return sum(nuc1 != nuc2 for nuc1, nuc2 in zip(seq1, seq2))


def run_flash(args):
    timestamp = datetime.now()
    destination, sample = args

    input_1, input_2 = sample["R1"], sample["R2"]
    prefix = "index%03i" % (sample["#"],)
    logfile = destination / f"{prefix}.stdout"

    command = ["flash", "-t", "1", "-M", "151", "-z", "-o", prefix, "-d", destination, input_1, input_2]

    log = logging.getLogger("flash")
    log.debug("Sample %03i: Merging reads into %s.*", sample["#"], destination / prefix)
    with open(logfile, "w") as loghandle:
        if run_command(command, stdout=loghandle):
            log.error("error while running FLASH; see log at %r", logfile)
            return False

    delta = (datetime.now() - timestamp).total_seconds()
    log.info("Sample %03i merged in %.1f seconds", sample["#"], delta)

    return True


def merge_fastq(destination, samples, threads):
    log = logging.getLogger("flash")
    log.info("Merging FASTQ files")

    args = [(destination, sample) for sample in samples]
    pool = Pool(processes=threads)
    for result in pool.imap(run_flash, args):
        if not result:
            pool.terminate()
            return False

    return True


def count_fastq_sequences(filename, min_length=MIN_READ_LENGTH):
    counts = defaultdict(int)

    with gzopen(filename, mode="rt") as handle:
        for idx, line in enumerate(handle):
            if idx % 4 == 1:
                read = line.strip().upper()
                if len(read) > min_length:
                    counts[read] += 1

    return dict(counts)


def collect_fastq_files(log, root, recursive):
    log.info(f"Locating FASTQ files in {root}")

    fastq_files = defaultdict(dict)
    for (dirpath, _, filenames) in os.walk(root, followlinks=True):
        for filename in filenames:
            if filename.endswith(READ_FORMAT):
                # someName_Si_L001_Rx_001.fastq.gz
                fields = filename.split("_")
                name = fields[0]
                sample = fields[-4]  # x = sequential sample numbering prefixed with 'S'
                pair = fields[-2]  # i = R1 or R2 (paired end)

                pairs = fastq_files[(int(sample[1:]), name)]
                if pair in pairs:
                    log.error("Multiple %s files for %s (%s)", pair, name, sample)
                    return None
                elif pair not in ("R1", "R2"):
                    log.error("Unexpected %s file for %s (%s)", pair, name, sample)
                    return None

                pairs[pair] = os.path.join(dirpath, filename)

        if not recursive:
            break

    samples = []
    for idx, ((sample, name), pairs) in enumerate(sorted(fastq_files.items()), start=1):
        if len(pairs) != 2:
            log.error("Missing files for sample %r (%i)", sample, name)
            return None

        samples.append({"#": idx, "name": name, "R1": pairs["R1"], "R2": pairs["R2"]})

    log.info("Identified %g FASTQ file pairs", len(samples))

    return samples


##############################################################################
# Reading, checking, and post-processing of FASTA target sequences


def read_fasta(fasta_file):
    name = None
    sequence = []
    records = []

    with open(fasta_file) as handle:
        for line in handle:
            if line.startswith(">"):
                if name is not None:
                    records.append((name, "".join(sequence)))

                # Leave out header meta-data, if present
                name = line[1:].split(None, 1)[0]
                sequence = []
            else:
                sequence.append(line.strip())

        if name is not None:
            records.append((name, "".join(sequence)))

    return records


def merge_identical_targets(targets):
    log = logging.getLogger("fasta")

    targets_by_barcodes = defaultdict(list)
    for target in targets:
        targets_by_barcodes[(target.barcode_5p, target.barcode_3p)].append(target)

    filtered_targets = []
    for group in targets_by_barcodes.values():
        if len(group) > 1:
            if any(group[0].seq != target.seq for target in group[1:]):
                log.error("Found targets with identical barcodes but different sequences:")
                for target in group:
                    log.error("    - %s", target.name)
                return None

            log.critical("Merging targets with identical sequence:")
            for target in group:
                log.critical("    - %s", target.name)

            name = "__OR__".join(sorted(set(target.name for target in group)))
            group = [group[0]._replace(name=name)]

        filtered_targets.extend(group)

    return filtered_targets


def compare_targets_5p(target1, target2, pos, max_ham, log):
    ham = hamming(target1.barcode_5p, target2.seq[pos : pos + BARSIZE])
    if ham < max_ham:
        log.warning(
            "Barcode from %r appears in %r, ham = %s", target1.name, target2.name, ham
        )
        log.warning(
            "    %s%s%s...",
            "." * pos,
            target1.barcode_5p,
            "." * (WORRY_ZONE - BARSIZE - pos),
        )
        log.warning("    %s...", target2.seq[:WORRY_ZONE])


def compare_targets_3p(target1, target2, pos, max_ham, log):
    offset = len(target2.seq) - WORRY_ZONE + pos
    ham = hamming(target1.barcode_3p, target2.seq[offset : offset + BARSIZE])
    if ham < max_ham:
        log.warning(
            "Barcode from %r appears in %r, ham = %s", target1.name, target2.name, ham
        )
        log.warning(
            "    ...%s%s%s",
            "." * pos,
            target1.barcode_3p,
            "." * (WORRY_ZONE - BARSIZE - pos),
        )
        log.warning("    ...%s", target2.seq[-WORRY_ZONE:])


def check_targets(targets, max_ham):
    # 1. Check for identical sequences and sequences with identical barcodes
    targets = merge_identical_targets(targets)
    if targets is None:
        return

    # 2. Check for FASTA records with the same name but different sequences
    log = logging.getLogger("fasta")
    targets_by_name = {}
    for target in targets:
        if target.name in targets_by_name:
            log.error("Multiple, different target sequences with the name %r", target.name)
            return

        targets_by_name[target.name] = target

    # 2: iterate each barcode through worry zone
    for key1, key2 in combinations(targets_by_name, 2):
        target1 = targets_by_name[key1]
        target2 = targets_by_name[key2]

        for pos in range(WORRY_ZONE - BARSIZE + 1):
            compare_targets_5p(target1, target2, pos, max_ham, log)
            compare_targets_5p(target2, target1, pos, max_ham, log)
            compare_targets_3p(target1, target2, pos, max_ham, log)
            compare_targets_3p(target2, target1, pos, max_ham, log)

        # get nearest ham
        ham_5p = hamming(target1.barcode_5p, target2.barcode_5p)
        ham_3p = hamming(target1.barcode_3p, target2.barcode_3p)

        # set nearest ham
        targets_by_name[target1.name] = target1._replace(ham_5p=min(target1.ham_5p, ham_5p), ham_3p=min(target1.ham_3p, ham_3p))
        targets_by_name[target2.name] = target2._replace(ham_5p=min(target2.ham_5p, ham_5p), ham_3p=min(target2.ham_3p, ham_3p))

    targets = sorted(targets_by_name.values(), key=lambda item: item.name.lower())

    name_len = max(len(x.name) for x in targets)
    log.info("{:<{}} {:>8} {:>8} {:>8}".format("Name", name_len, "Length", "5p Ham", "3p Ham"))
    for target in targets:
        log.info("{:<{}} {:>8} {:>8} {:>8}".format(target.name, name_len, target.len, target.ham_5p, target.ham_3p))

    return targets


def build_targets(fasta, max_ham):
    targets = []
    for (name, seq) in fasta:
        seq = seq.upper()
        targets.append(TargetSequence(
            name=name, seq=seq, len=len(seq), barcode_5p=seq[:BARSIZE], barcode_3p=seq[-BARSIZE:],
            ham_5p=max_ham, ham_3p=max_ham,
        ))

    return check_targets(targets, max_ham)


##############################################################################


def strict_matching(targets, read_counts, results, forlog):
    for target in targets:
        for read, count in tuple(read_counts.items()):
            if read.startswith(target.barcode_5p) and read.endswith(target.barcode_3p):
                results[target.name][len(read) - target.len] += count
                read_counts.pop(read)  # remove analyzed reads

                forlog[target.name][0] += count


def relaxed_matching(args, targets, read_counts, results, forlog):
    for target in targets:
        # long = l = BARSIZE, short = s = REDUCED_BAR
        # relaxed = r = ham < nearest ham, ultra relaxed = u = ham < max_ham
        # We never use "short short" due to risk of off-targets, but "ultru
        # ultra" (uncertain origin) will be done in 'ultra relaxed matching' which is
        # very slow

        # lbar, lham, rbar, rham
        parameters = (
            # 1. lrsu: 5' long relaxed, 3' short ultra
            (BARSIZE, target.ham_5p, REDUCED_BAR, args.hamming_distance),
            # 2. sulr: 5' short ultra, 3' long relaxed
            (REDUCED_BAR, args.hamming_distance, BARSIZE, target.ham_3p),
            # 3. srlu: 5' short relaxed, 5' long ultra
            (REDUCED_BAR, target.ham_5p, BARSIZE, args.hamming_distance),
            # 4. lusr: 5' long ultra, 5' short relaxed
            (BARSIZE, args.hamming_distance, REDUCED_BAR, target.ham_3p),
        )

        for idx, (lbar, lham, rbar, rham) in enumerate(parameters, start=1):
            for read, count in tuple(read_counts.items()):
                if (
                    hamming(read[:lbar], target.barcode_5p[:lbar]) < lham
                    and hamming(read[-rbar:], target.barcode_3p[-rbar:]) < rham
                ):
                    results[target.name][len(read) - target.len] += count
                    forlog[target.name][idx] += count
                    read_counts.pop(read)


def ultra_relaxed_matching(args, targets, read_counts, results, forlog):
    for read, count in tuple(read_counts.items()):
        # find out where a read MIGHT originate from
        matching_pcrs = []
        min_score = args.hamming_distance
        for target in targets:
            score = hamming(
                target.barcode_5p[:REDUCED_BAR], read[:REDUCED_BAR]
            ) + hamming(target.barcode_3p[-REDUCED_BAR:], read[-REDUCED_BAR:])
            if score < min_score:
                matching_pcrs = [target]
            elif score == min_score:
                matching_pcrs.append(target)

        if matching_pcrs:
            # ok this read might be something.
            # check lusu or sulu
            maybe_product = set()
            for target in matching_pcrs:
                # lusu
                if (
                    hamming(target.barcode_5p, read[:BARSIZE]) < args.hamming_distance
                    and hamming(target.barcode_3p[-REDUCED_BAR:], read[-REDUCED_BAR:])
                    < args.hamming_distance
                ):
                    maybe_product.add(target)
                # sulu
                if (
                    hamming(target.barcode_5p[:REDUCED_BAR], read[:REDUCED_BAR])
                    < args.hamming_distance
                    and hamming(target.barcode_3p, read[-BARSIZE:])
                    < args.hamming_distance
                ):
                    maybe_product.add(target)

            if len(maybe_product) == 1:
                target = maybe_product.pop()
                results[target.name][target.len - len(read)] += count
                read_counts.pop(read)  # remove analyzed reads
                forlog[target.name][5] += count
            elif len(maybe_product) > 1:
                min_score = args.hamming_distance * 2
                winner = []
                for target in maybe_product:
                    score = hamming(target.barcode_5p, read[:BARSIZE]) + hamming(
                        target.barcode_3p, read[-BARSIZE:]
                    )
                    if score < min_score:
                        min_score = score
                        winner = [target]
                    elif score == min_score:
                        winner.append(target)

                if len(winner) == 1:
                    target = winner[0]
                    results[target.name][target.len - len(read)] += count
                    read_counts.pop(read)  # remove analyzed reads
                    forlog[target.name][5] += count
                elif len(winner) > 1:
                    log = logging.getLogger("matching")
                    log.warning("ambigious reads:\n%s", read)
                    for target in winner:
                        log.warning(target.name)


def analyze_sample(parameters):
    args, sample, targets = parameters
    log = logging.getLogger("matching")

    results = {}
    for target in targets:
        results[target.name] = defaultdict(int)

    filename = os.path.join(args.output_merged, "index%03i.extendedFrags.fastq.gz" % (sample["#"],))
    log.debug("Sample %03i: Reading index from %s", sample["#"], filename)
    read_counts = count_fastq_sequences(filename)
    total_reads = sum(read_counts.values())

    logstats = {}
    for target in targets:
        # strict, lrsu, sulr, srlu, lusr, ur total
        logstats[target.name] = [0, 0, 0, 0, 0, 0]

    # STRICT matching is very fast and assigns anything we can easily identify
    log.debug("Sample %03i: Strict matching %i products with %i reads", sample["#"], len(read_counts), sum(read_counts.values()),)
    strict_matching(targets, read_counts, results, logstats)

    if not args.deactivate_relaxed_matching:
        # RELAXED matching (slower, but not very)
        log.debug("Sample %03i: Relaxed matching %i products with %i reads", sample["#"], len(read_counts), sum(read_counts.values()),)
        relaxed_matching(args, targets, read_counts, results, logstats)

        # ULTRA RELAXED MATCHING (probably slow)
        # now anything left will be ultra relaxed matched. This opens up for a read to
        # match multiple targets, so we need to check for this and also any off-targets
        # these primers may be producing will also show up here...maybe that should be
        # a number to report...
        log.debug("Sample %03i: Ultra relaxed matching %i products with %i reads", sample["#"], len(read_counts), sum(read_counts.values()),)
        ultra_relaxed_matching(args, targets, read_counts, results, logstats)

    return {"#": sample["#"], "name": sample["name"], "#reads": total_reads, "#unmatched": sum(read_counts.values()), "targets": results, "logstats": logstats, "unmatched": read_counts}


def log_sample_stats(sample):
    log = logging.getLogger("matching")

    name_length = max(len(x) for x in sample["logstats"])
    log.info(
        "{:<{}} {:>8} {:>8} {:>8} {:>8} {:>8} {:>8} {:>8}".format(
            "Target", name_length, "Total", "Strict", "lrsu", "sulr", "srlu", "lusr", "ur",
        )
    )
    for header, numbers in sample["logstats"].items():
        log.info(
            "{:<{}} {:>8} {:>8} {:>8} {:>8} {:>8} {:>8} {:>8}".format(
                header, name_length, sum(numbers), *numbers
            )
        )

    unmatched_reads = []
    for read, count in sample["unmatched"].items():
        if (
            count >= SIGNIFICANT_UNMATCHED_THRESHOLD_READS
            and len(read) >= SIGNIFICANT_UNMATCHED_THRESHOLD_LENGTH
        ):
            unmatched_reads.append((count, read))

    if unmatched_reads:
        log.warning("Unmatched reads, index: %s", sample["#"])
        for count, read in sorted(unmatched_reads):
            log.warning(" % 6i: %s", count, read)

    log.info("Unmatched unique reads: %s", len(sample["unmatched"]))
    log.info("Unmatched total reads: %s", sum(sample["unmatched"].values()))


def analyze(args, samples, fasta):
    log = logging.getLogger("matching")

    results = []  # results are [{name: defaultdict(int), ...}, ...]
    pool = Pool(processes=args.t)
    pool_args = [(args, sample, fasta) for sample in samples]
    for output in pool.imap(analyze_sample, pool_args):
        log.info(f"Completed analysis of sample {output['#']} ..")
        log_sample_stats(sample=output)

        output.pop("unmatched")
        results.append(output)

    if results:
        output_summary = {"#": "*", "name": "summary", "logstats": results[0].pop("logstats"), "unmatched": {}, }

        for result in results[1:]:
            for name, numbers in result.pop("logstats").items():
                for idx, count in enumerate(numbers):
                    output_summary["logstats"][name][idx] += count

        log.info(f"Completed analysis of {len(results)} samples!")
        log_sample_stats(sample=output_summary)

    return results


def write_report(args, results, output_filename):
    workbook = Workbook()

    worksheet1 = workbook.active
    worksheet1.title = "results"
    worksheet1.cell(1, 1, "Index/Well")
    worksheet1.cell(1, 2, "Sample")
    worksheet1.cell(1, 3, "Target")
    worksheet1.cell(1, 4, "Reads")
    worksheet1.cell(1, 5, "WT")
    worksheet1.cell(1, 6, "Indel")
    worksheet1.cell(1, 7, "%Indel")
    # peak counting
    max_peaks = 0
    for sample in results:
        for target, peaks in sample["targets"].items():
            sigpeaks = len([p for s, p in peaks.items() if s and p >= TARGET_LOCAL_THRESHOLD_READS and (float(p) / sum(peaks.values())) >= TARGET_LOCAL_THRESHOLD_PERCENTAGE])
            if sigpeaks > max_peaks:
                max_peaks = sigpeaks
    # add a header per peak
    peak_col_start = 7
    for col in range(peak_col_start, max_peaks + peak_col_start):
        worksheet1.cell(1, 1+col, "peak{}".format(col - (peak_col_start - 1)))
        worksheet1.cell(1, 1+col + max_peaks, "peak{}%".format(col - (peak_col_start - 1)))
    row = 1
    for sample in results:
        for target, peaks in sample["targets"].items():
            if sum(peaks.values()) >= TARGET_TOTAL_THRESHOLD:
                if args.convert_to_wells:
                    rows, columns = args.plate_layout
                    well = convert_index_to_well(sample["#"], rows, columns)

                    worksheet1.cell(row+1, 1, well)
                else:
                    worksheet1.cell(row+1, 1, sample["#"])
                worksheet1.cell(row+1, 2, sample["name"])
                worksheet1.cell(row+1, 3, target)
                worksheet1.cell(row+1, 4, sum(peaks.values())).style = 'Comma [0]'
                worksheet1.cell(row+1, 5, peaks[0]).style = "Comma [0]"
                worksheet1.cell(row+1, 6, sum(peaks.values()) - peaks[0]).style = "Comma [0]"
                worksheet1.cell(row+1, 7, "=F{0}/D{0}".format(row + 1)).style = 'Percent'
                column = 7
                for indel_size, num in sorted(peaks.items()):
                    if indel_size:
                        if (
                            num > TARGET_LOCAL_THRESHOLD_READS
                            and (float(num) / sum(peaks.values()))
                            >= TARGET_LOCAL_THRESHOLD_PERCENTAGE
                        ):
                            if indel_size % 3 == 0:
                                indel_size = "{} (inframe)".format(indel_size)
                            worksheet1.cell(row+1, column+1, "{}".format(indel_size))
                            column += 1
                            worksheet1.cell(row+1, column + max_peaks, f"={num}/D{row + 1}").style = 'Percent'
                row += 1
    worksheet1.auto_filter.ref = worksheet1.dimensions

    for cs in worksheet1.columns:
        worksheet1.column_dimensions[cs[0].column_letter].bestFit = True

    indel_sizes = []
    for i in results:
        q = pd.DataFrame(i['targets']).T
        q['index'] = i['#']
        indel_sizes.append(q)
    indel_df = pd.concat(indel_sizes).set_index('index', append=True, drop=True).rename_axis(('target', 'index')).sort_index(axis=1).reset_index()
    worksheet2 = workbook.create_sheet("indel_sizes")
    for r in dataframe_to_rows(indel_df, index=False, header=True):
        worksheet2.append(r)
    worksheet2.auto_filter.ref = worksheet2.dimensions
    workbook.save(output_filename)


def write_json(args, targets, results, output_filename):
    data = {
        "samples": {},
        "targets": {target.name: target.seq for target in targets},
        "version": __version__,
        "settings": {
            "-cw": args.convert_to_wells,
            "-dr": args.deactivate_relaxed_matching,
            "-hd": args.hamming_distance,
            "-pl": args.plate_layout,
        },
    }

    for sample in results:
        result = {}
        for target_name, target_data in sample["targets"].items():
            peaks = dict(target_data)
            total_reads = sum(peaks.values())
            total_wildtype = peaks.pop(0, 0)
            total_indels = total_reads - total_wildtype

            if total_reads:
                result[target_name] = {
                    "reads": total_reads,
                    "reads_wildtype": total_wildtype,
                    "reads_mutant": total_indels,
                    "peaks": peaks,
                }

        data["samples"][sample["#"]] = {
            "name": sample["name"],
            "reads": sample["#reads"],
            "unidentified": sample["#unmatched"],
            "targets": result,
        }

    with open(output_filename, "wt") as handle:
        handle.write(json_dumps(data))


def parse_args():
    parser = ArgumentParser(
        description="hamplicons - estimating indel sizes in amplicons using Hamming distances",
        epilog="Needs flash available in PATH " "(http://ccb.jhu.edu/software/FLASH/)",
        formatter_class=ArgumentDefaultsHelpFormatter,
    )

    parser.add_argument('-v', action="version", version="%(prog)s v" + __version__)
    parser.add_argument('-t', type=int, help="Number of threads for read merging/analysis", default=cpu_count(), metavar='N')
    parser.add_argument("--log-level", type=str.upper, default="INFO", choices=("DEBUG", "INFO", "WARNING", "ERROR"), help="Log-level of messages printed STDERR and written to hamp_out.log")

    group = parser.add_argument_group("Input")
    group.add_argument("-tf", "--targets-fasta", type=str, default="targets.fa", help="Fasta file containing WT PCR sequences", metavar="TARGETS",)
    group.add_argument("-dd", "--data-directory", type=str, default="fastq", help="Directory containing fastq.gz files", metavar="DIRECTORY",)
    group.add_argument("-ddr", "--data-directory-recursive", action="store_true", help="Search data directory recursively", )
    group.add_argument("-sm", "--skip-merging", action="store_true", help="skip merging of fastq paired end reads",)

    group = parser.add_argument_group("Matching")
    group.add_argument("-dr", "--deactivate-relaxed-matching", action="store_true", help="Deactivate relaxed matching",)
    group.add_argument("-hd", "--hamming-distance", type=int, help="Allow this many mismatches in relaxed matching", default=4, metavar="INT",)

    group = parser.add_argument_group("Output")
    group.add_argument('-o', default="output", help="Folder to put all output in", type=Path, metavar="output_folder")
    group.add_argument("-of", "--output-formats", nargs="+", choices=("json", "xlsx"), metavar="FORMAT", default="xlsx", help="One or more formats (xlsx or json) in which to save analysis results",)
    group.add_argument("-cw", "--convert-to-wells", action="store_true", help="Convert Index to Well", )
    group.add_argument("-pl", "--plate-layout", type=list, nargs=2, help="Define layout for index translation when using --convert-to-wells. " "E.g. `-pl 8 12` for 8 rows and 12 columns", default=[8, 12], metavar="INT",)

    args = parser.parse_args()
    # check existence of input files
    if not Path(args.targets_fasta).is_file():
        parser.print_help()
        sys.exit()
    return args


def main(args):
    log = setup_logging(level=args.log_level)

    # if we are merging, check that flash exists
    if not (args.skip_merging or which("flash")):
        log.error("Required executable `flash` not found in PATH!")
        log.error("Please install FLASH from 'http://ccb.jhu.edu/software/FLASH/'")

        return 1

    args.output_merged = args.o / "hamp_out.merged"
    args.output_merged.mkdir(parents=True, exist_ok=True)

    setup_file_logging(filename=args.o / "hamp_out.log", level=args.log_level)

    log.info(f"Reading FASTA file {args.targets_fasta}")
    targets = build_targets(read_fasta(args.targets_fasta), max_ham=args.hamming_distance)

    if targets is None:
        return 1

    samples = collect_fastq_files(log, args.data_directory,recursive=args.data_directory_recursive,)

    if samples is None:
        return 1

    if args.skip_merging:
        log.info("Skipping FASTQ merging")
    elif not merge_fastq(args.output_merged, samples, args.t):
        return 1

    log.info("Analyzing FASTQ files")
    results = analyze(args, samples, targets)

    if "xlsx" in args.output_formats:
        log.info(f"Writing report to {args.o / 'hamp_out.xlsx'}")
        write_report(args=args, results=results, output_filename=args.o / "hamp_out.xlsx",)

    if "json" in args.output_formats:
        log.info(f"Writing JSON to {args.o / 'hamp_out.json'}")
        write_json(args=args, targets=targets, results=results, output_filename=args.o / "hamp_out.json",)

    log.info("Done")


def entry_point():
    sys.exit(main(parse_args()))


if __name__ == "__main__":
    entry_point()
